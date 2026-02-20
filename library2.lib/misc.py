#! python3
# -*- coding: utf-8 -*-
from System.Collections.Generic import List

import os
from datetime import datetime

import openpyxl

from Autodesk.Revit.DB import (
    Transaction,
    ElementId,
    BuiltInCategory,
    ExternalDefinitionCreationOptions,
    SpecTypeId,
    FilteredElementCollector,
    UnitUtils,
    UnitTypeId,
    FilteredWorksetCollector,
    WorksetKind
)
from Autodesk.Revit.UI import TaskDialogResult
from debug import debug_print
from material import materiaux, unknown_materials

try:
    from Autodesk.Revit.DB import GroupTypeId
    PG_DATA = GroupTypeId.Data
except:
    from Autodesk.Revit.DB import BuiltInParameterGroup
    PG_DATA = BuiltInParameterGroup.PG_DATA

from error import OperationResult, ErrorCode
from forms import show_taskdialog
from paths import D_MODEL_DIRECTORY_PATH
from views import three_d_and_isolate


def convert_to_cubic_meters(volume_in_cubic_feet):
    """
    converts cubic feet to cubic meters

    Parameters
    ----------
    volume_in_cubic_feet : int
        to be converted to cubic meters.

    Returns
    -------
    int
        number after conversion from cubic feet to cubic meters.

    """
    volume_in_cubic_meters = UnitUtils.ConvertFromInternalUnits(volume_in_cubic_feet, UnitTypeId.CubicMeters)
    return volume_in_cubic_meters


def remove_duplicate_dicts(lst):
    """
    removes duplicate dictionaries from param lst

    Parameters
    ----------
    lst : list
        a list in need of duplicate removal.

    Returns
    -------
    unique_dicts : a list with duplicates removed.

    """
    unique_dicts_list = []
    seen_dicts = set()
    for d in lst:
        dict_items = tuple(sorted(d.items()))
        if dict_items not in seen_dicts:
            unique_dicts_list.append(d)
            seen_dicts.add(dict_items)

    return unique_dicts_list


def create_project_parameter(document):
    """
    Create a project parameter in a Revit document.

    Returns:
        OperationResult: Success with message or error with details
    """
    t = Transaction(document, "Create Project Parameter")

    try:
        t.Start()

        app = document.Application
        shared_parameter_file = app.OpenSharedParameterFile()

        # Check if shared parameter file is set
        if shared_parameter_file is None:
            t.RollBack()
            return OperationResult.error_result(
                ErrorCode.SHARED_PARAMETER_FILE_NOT_SET,
                message='Aucun fichier de paramètres partagés n\'est configuré.'
            )

        # Check if EMC2B group exists
        try:
            emc2b_group = shared_parameter_file.Groups.get_Item("EMC2B")
            if emc2b_group is None:
                t.RollBack()
                return OperationResult.error_result(
                    ErrorCode.SHARED_PARAMETER_FILE_NOT_SET,
                    message='Le groupe "EMC2B" est introuvable.'
                )
        except Exception:
            t.RollBack()
            return OperationResult.error_result(
                ErrorCode.SHARED_PARAMETER_FILE_NOT_SET
            )

        param_name = "Project_Id"
        param_forge_type = SpecTypeId.String.Text
        options = ExternalDefinitionCreationOptions(param_name, param_forge_type)

        definition = None
        for d in shared_parameter_file.Groups.get_Item("EMC2B").Definitions:
            if d.Name == param_name:
                definition = d
                break

        if definition is None:
            group = shared_parameter_file.Groups.get_Item("EMC2B")
            definition = group.Definitions.Create(options)

        categories = document.Settings.Categories

        target_category = None
        for cat in categories:
            if cat.Name == "Informations sur le projet":
                target_category = cat
                break
        if target_category is None:
            raise Exception("La catégorie 'Informations sur le projet' n'a pas été trouvée.")

        category_set = app.Create.NewCategorySet()
        category_set.Insert(target_category)

        instance_binding = app.Create.NewInstanceBinding(category_set)
        document.ParameterBindings.Insert(
            definition, instance_binding, PG_DATA
        )

        t.Commit()
        return OperationResult.success_result("Parameter created successfully.")

    except Exception as e:
        if t.HasStarted():
            t.RollBack()
        return OperationResult.error_result(
            ErrorCode.TRANSACTION_FAILED,
            details=str(e)
        )

def assign_project_id_value(document, results):
    project_parameter_name = "Project_Id"

    transaction = Transaction(document, "Set Shared Parameter Value")
    transaction.Start()

    collector = FilteredElementCollector(document)
    project_info = (
        collector.OfCategory(BuiltInCategory.OST_ProjectInformation).WhereElementIsNotElementType().ToElements())

    for info in project_info:
        project_shared_param = info.LookupParameter(project_parameter_name)

        if project_shared_param:
            project_shared_param.Set(results["project_id"])

    transaction.Commit()


def export_elements_to_excel(element_list, export_filepath):
    from openpyxl.styles import Font, PatternFill, Alignment

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    headers = ["Id de l'élément", "Nature", "Elément", "Catégorie", "Famille", "Classe", "Volume"]
    worksheet.append(headers)

    # Format header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add data rows
    for row in element_list:
        worksheet.append(row)

    # Auto-adjust column widths
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value)) if cell.value is not None else 0
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)

    workbook.save(export_filepath)


def export_to_excel(no_material):
    """
    Export elements without material to an Excel file.

    :param no_material: list of elements without material.
    :return: Path to the exported Excel file.
    """
    timestamp_string = datetime.now().strftime("%Y-%m-%d_%H.%M.%S")
    excel_filename = "Objets sans matériaux_{}.xlsx".format(timestamp_string)
    excel_filepath = os.path.join(D_MODEL_DIRECTORY_PATH, excel_filename)

    try:
        export_elements_to_excel(no_material, excel_filepath)
        return excel_filepath
    except Exception as export_error:
        show_taskdialog(
            title="Erreur",
            main_instruction="Échec de l'exportation Excel.",
            main_content=str(export_error)
        )
        raise


def process_and_export_to_excel(ui_document, elements_list, show_results=False, create_view=False):
    document = ui_document.Document
    elements_without_material_rows = []
    element_ids_to_isolate = List[ElementId]()

    for element in elements_list:
        materials_data_rows, element_id = process_element_materials(element, document)

        if not materials_data_rows:
            continue

        elements_without_material_rows.extend(materials_data_rows)

        if element_ids_to_isolate is not None and element_id:
            element_ids_to_isolate.Add(element_id)

    if not elements_without_material_rows:
        if show_results:
            show_taskdialog(
                title="Objets sans matériaux",
                main_instruction="Votre modèle est super !",
                main_content="Vous n'avez aucun élément sans matériaux."
            )
        return

    excel_filepath = export_to_excel(elements_without_material_rows)

    if show_results:
        display_results(len(elements_without_material_rows), excel_filepath)

    if create_view and element_ids_to_isolate:
        three_d_and_isolate(ui_document, element_ids_to_isolate)


def worksharing_check(document):
    """Vérifier si le document est partagé et si les sous-projets AREP y sont"""
    if not document.IsWorkshared:
        return OperationResult.error_result(
            ErrorCode.NOT_WORKSHARED,
            details="Veuillez partager la maquette et relancer le calcul.\n"
                    "Les sous-projets acceptés doivent commencer par 2_, 3_ ou 4_ pour être reconnus.",
            context='worksharing_check'
        )

    workset_list = FilteredWorksetCollector(document).OfKind(WorksetKind.UserWorkset)

    valid_worksets = [ws for ws in workset_list if ws.Name.startswith(('2_', '3_', '4_'))]

    if not valid_worksets:
        return OperationResult.error_result(
            ErrorCode.NO_VALID_WORKSETS,
            message="La maquette est partagée mais aucun sous-projet AREP valide n'a été trouvé. "
                    "Merci de vérifier la présence de ces sous-projets et relancer le calcul.",
            details="Dans la version actuelle, les sous-projets acceptés doivent commencer par 2_, 3_ ou 4_ pour être reconnus.",
            context='worksharing_check'
        )

    return OperationResult.success_result({
        'valid_worksets': valid_worksets,
        'total_worksets': len(list(workset_list))
    })


def get_element_id_value(element_id):
    """
    Safely gets the integer value of an ElementId, handling different Revit API versions.

    Parameters
    ----------
    element_id : ElementId
        The ElementId to get the integer value from.
    Returns
    -------
    int
        The integer value of the ElementId.
    """
    return element_id.Value if hasattr(element_id, 'Value') else element_id.IntegerValue


def get_element_volume(element):
    try:
        return convert_to_cubic_meters(element.LookupParameter("Volume").AsDouble())
    except:
        return ""


def has_unknown_dependent_materials(element, doc):
    """Check if element has dependent elements with unknown materials."""
    debug_print("----Dependent_elements")
    for dep_id in element.GetDependentElements(None):
        dep_materials = doc.GetElement(dep_id).GetMaterialIds(False)
        if dep_materials:
            for mat_id in dep_materials:
                mat = doc.GetElement(mat_id)
                if element.GetMaterialVolume(mat_id):
                    if not materiaux.is_known_material(mat.MaterialClass):
                        debug_print("----------Matériau inconnu : {}".format(mat.MaterialClass))
                        return True
    return False


def process_element_materials(element, doc):
    element_id = element.Id
    famille = element.LookupParameter("Famille").AsValueString() # TODO: find a non language specific way
    volume = get_element_volume(element)

    debug_print("**{} ** {}".format(element_id, famille))

    if not has_unknown_dependent_materials(element, doc):
        return None, None

    materials = element.GetMaterialIds(False)

    if not materials:
        return [[
            str(element_id),
            "Elément",
            str(element.Name),
            str(element.Category.Name),
            str(famille),
            "",
            volume
        ]], element_id

    unknown_materials_data = []
    debug_print("----Materials")

    count = 0
    for mat_id in materials:
        mat = doc.GetElement(mat_id)
        debug_print("-------{} -- {}".format(mat.Name, mat.MaterialClass))
        if element.GetMaterialVolume(mat_id):
            if mat.MaterialClass in unknown_materials:
                id_label = "╰----->" if count else str(element_id)
                unknown_materials_data.append([
                    id_label,
                    "Matériau",
                    str(mat.Name),
                    str(famille),
                    str(element.Category.Name),
                    str(mat.MaterialClass),
                    volume
                ])
                count += 1

    if unknown_materials_data:
        return unknown_materials_data, element_id

    return None, None


def display_results(element_count, excel_filepath):
    plural_suffix = "s" if element_count > 1 else ""

    main_instruction = "{} élément{} sans matériaux ont été trouvés.".format(element_count, plural_suffix)
    main_content = (
        "Une vue 3D contenant les éléments sans matériaux (ou avec des matériaux inconnus) sera créée.\n\n"
        "Un export Excel contenant le détail de ces éléments a été enregistré à cet emplacement :\n"
        "\"{}\"\n\n"
        "Voulez-vous ouvrir le fichier Excel maintenant ?".format(D_MODEL_DIRECTORY_PATH.replace("\\", "/"))
    )

    user_choice = show_taskdialog(
        title="Éléments sans matériaux",
        main_instruction=main_instruction,
        main_content=main_content,
        buttons=['Yes', 'No']
    )

    if user_choice == TaskDialogResult.Yes:
        os.startfile(excel_filepath)
